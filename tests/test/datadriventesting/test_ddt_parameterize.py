import openpyxl
import pytest

from src.constants.api_constants import APIConstants
from src.utils.utils import Util
from src.helpers.api_requests_wrapper import *
from src.helpers.common_verification import *

def read_credentials_from_excel(file_path):
    credentials = []
    workbook = openpyxl.load_workbook(filename=file_path) #The load_workbook() function opens file for reading
    sheet1 = workbook.active #The object of the workbook.active has been created in the script to read the values of the max_row and the max_column properties.
    #for loop to read each row from the excel sheet
    for row in sheet1.iter_rows(min_row=2,values_only=True):
        username,password = row
        credentials.append({"username":username,
                            "password":password})
    return credentials

def create_auth_request(username, password):
        payload = {"username": username,
                   "password": password}
        response = post_request(
            url=APIConstants.url_create_token(),
            headers=Util().common_headers_json(),
            auth=None,
            payload=payload,
            in_json=False
        )
        return response
#instead of for loop to read each row from excel sheet, we use parametrize to all data(each row)
#Parameterizing of a test is done to run the test against multiple sets of inputs.
@pytest.mark.parametrize("user_cred", read_credentials_from_excel("C:\All\Python_Automation\Py2xAPIAutomationFramework\\tests\\test\datadriventesting\\testdata_ddt_123.xlsx"))
def test_create_auth_with_excel(user_cred):
    username = user_cred["username"]
    password = user_cred["password"]
    print(username, password)
    response = create_auth_request(username=username, password=password)
    print(response.status_code)
    verify_http_status_code(response_data=response, expect_data=200)
